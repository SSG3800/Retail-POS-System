import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import customtkinter as ctk
import sqlite3
import os
import hashlib
from openpyxl import Workbook
from datetime import datetime
from PIL import Image, ImageTk
import win32api
import win32print

# Set appearance mode and color theme
ctk.set_appearance_mode("Light")
ctk.set_default_color_theme("blue")


class PasswordDialog(ctk.CTkToplevel):
    """A dialog for password entry."""

    def __init__(self, parent):
        super().__init__(parent)
        self.title("Enter Password")
        self.transient(parent)
        self.grab_set()
        self.password_ok = False
        self.label = ctk.CTkLabel(self, text="Please enter your password to proceed:")
        self.label.pack(padx=20, pady=(20, 10))
        self.password_entry = ctk.CTkEntry(self, show="*")
        self.password_entry.pack(padx=20, pady=10, fill="x")
        self.password_entry.focus()
        self.ok_button = ctk.CTkButton(self, text="OK", command=self.on_ok)
        self.ok_button.pack(side="left", padx=(20, 10), pady=20, expand=True)
        self.cancel_button = ctk.CTkButton(self, text="Cancel", command=self.on_cancel)
        self.cancel_button.pack(side="right", padx=(10, 20), pady=20, expand=True)
        self.password_entry.bind("<Return>", self.on_ok)

    def on_ok(self, event=None):
        if self.master.verify_password(self.password_entry.get()):
            self.password_ok = True
            self.destroy()
        else:
            messagebox.showerror("Error", "Incorrect password.", parent=self)

    def on_cancel(self):
        self.password_ok = False
        self.destroy()

    def show(self):
        self.wait_window()
        return self.password_ok


class RetailApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Retail POS & Inventory")
        self.geometry("1400x800")
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.last_sale_details = None

        # --- Database Setup ---
        app_data_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
        os.makedirs(app_data_path, exist_ok=True)
        self.db_path = os.path.join(app_data_path, "retail.db")
        self.conn = sqlite3.connect(self.db_path)
        self.cursor = self.conn.cursor()
        self.create_tables()

        # --- Configure Styles ---
        style = ttk.Style()
        style.theme_use("default")
        style.configure("Treeview.Heading", font=('Arial', 14, 'bold'), background="yellow", foreground="black")
        style.configure("Treeview", rowheight=30, font=('Arial', 12))
        style.map('Treeview.Heading', background=[('active', '#FFD700')])  # Keep yellow on click

        # --- Sidebar ---
        self.navigation_frame = ctk.CTkFrame(self, corner_radius=0)
        self.navigation_frame.grid(row=0, column=0, sticky="nsew")
        self.navigation_frame.grid_rowconfigure(6, weight=1)
        self.navigation_frame_label = ctk.CTkLabel(self.navigation_frame, text="SAMARA trade center",
                                                   font=ctk.CTkFont(size=20, weight="bold"))
        self.navigation_frame_label.grid(row=0, column=0, padx=20, pady=20)
        self.inventory_button = self.create_nav_button("Inventory", self.inventory_button_event, 1)
        self.pos_button = self.create_nav_button("POS", self.pos_button_event, 2)
        self.sales_button = self.create_nav_button("Sales", self.sales_button_event, 3)
        self.export_button = self.create_nav_button("Export to Excel", self.export_to_excel_secure, 4)
        self.settings_button = self.create_nav_button("Settings", self.settings_button_event, 5)

        # --- Main Frames ---
        self.inventory_frame = ctk.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.pos_frame = ctk.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.sales_frame = ctk.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.settings_frame = ctk.CTkFrame(self, corner_radius=0, fg_color="transparent")

        self.create_inventory_ui()
        self.create_pos_ui()
        self.create_sales_ui()
        self.create_settings_ui()
        self.select_frame_by_name("inventory")

        # --- Global Key Bindings for Navigation ---
        self.bind_all("<Key>", self.handle_key_press)

    def handle_key_press(self, event):
        """Main handler for keyboard navigation."""
        focus = self.focus_get()
        if not focus: return

        if event.keysym == 'Return':
            if hasattr(focus, 'invoke'):
                focus.invoke()
        elif event.keysym == 'Up':
            focus.tk_focusPrev().focus()
        elif event.keysym == 'Down':
            focus.tk_focusNext().focus()
        elif event.keysym == 'Left':
            focus.tk_focusPrev().focus()
        elif event.keysym == 'Right':
            focus.tk_focusNext().focus()

    def create_nav_button(self, text, command, row):
        button = ctk.CTkButton(self.navigation_frame, corner_radius=0, height=40, text=text, font=ctk.CTkFont(size=14),
                               fg_color="transparent", text_color=("gray10", "gray90"),
                               hover_color=("gray70", "gray30"),
                               command=command)
        button.grid(row=row, column=0, sticky="ew")
        return button

    def hash_password(self, password):
        return hashlib.sha256(password.encode()).hexdigest()

    def verify_password(self, entered_password):
        self.cursor.execute("SELECT password FROM settings WHERE id = 1")
        stored_hash = self.cursor.fetchone()
        return stored_hash[0] == self.hash_password(entered_password) if stored_hash else False

    def ask_password(self):
        return PasswordDialog(self).show()

    def create_tables(self):
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS products (
                id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL,
                price REAL NOT NULL, quantity INTEGER NOT NULL,
                image_path TEXT DEFAULT '')''')
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS sales (
                id INTEGER PRIMARY KEY AUTOINCREMENT, total_price REAL NOT NULL,
                sale_date TIMESTAMP DEFAULT (datetime('now', 'localtime')))''')
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS sale_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT, sale_id INTEGER, product_id INTEGER,
                product_name TEXT, quantity INTEGER, price REAL,
                FOREIGN KEY (sale_id) REFERENCES sales (id),
                FOREIGN KEY (product_id) REFERENCES products (id))''')
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS settings (
                id INTEGER PRIMARY KEY CHECK (id = 1), password TEXT NOT NULL)''')
        self.cursor.execute("SELECT * FROM settings WHERE id = 1")
        if not self.cursor.fetchone():
            self.cursor.execute("INSERT INTO settings (id, password) VALUES (1, ?)", (self.hash_password("admin"),))
        self.conn.commit()

    # --- Inventory Section ---
    def create_inventory_ui(self):
        form_frame = ctk.CTkFrame(self.inventory_frame)
        form_frame.pack(fill="x", padx=20, pady=20)
        form_frame.grid_columnconfigure(2, weight=1)
        ctk.CTkLabel(form_frame, text="Add/Update Product", font=ctk.CTkFont(size=18, weight="bold")).grid(row=0,
                                                                                                           column=0,
                                                                                                           columnspan=2,
                                                                                                           padx=10,
                                                                                                           pady=10,
                                                                                                           sticky="w")
        self.product_id_entry = ctk.CTkEntry(form_frame, placeholder_text="Product ID (for updating)")
        self.product_id_entry.grid(row=1, column=0, padx=10, pady=5, sticky="ew")
        self.product_name_entry = ctk.CTkEntry(form_frame, placeholder_text="Name")
        self.product_name_entry.grid(row=1, column=1, padx=10, pady=5, sticky="ew")
        self.product_price_entry = ctk.CTkEntry(form_frame, placeholder_text="Price")
        self.product_price_entry.grid(row=2, column=0, padx=10, pady=5, sticky="ew")
        self.product_qty_entry = ctk.CTkEntry(form_frame, placeholder_text="Quantity")
        self.product_qty_entry.grid(row=2, column=1, padx=10, pady=5, sticky="ew")
        ctk.CTkButton(form_frame, text="Add Image", command=self.add_image).grid(row=3, column=0, padx=10, pady=10,
                                                                                 sticky="ew")
        self.image_path_label = ctk.CTkLabel(form_frame, text="No image selected")
        self.image_path_label.grid(row=3, column=1, padx=10, pady=10, sticky="w")
        self.product_image_label = ctk.CTkLabel(form_frame, text="")
        self.product_image_label.grid(row=1, column=2, rowspan=3, padx=20)
        ctk.CTkButton(form_frame, text="Add Product", command=self.add_product_secure).grid(row=4, column=0, padx=10,
                                                                                            pady=10, sticky="ew")
        ctk.CTkButton(form_frame, text="Update Product", command=self.update_product_secure).grid(row=4, column=1,
                                                                                                  padx=10, pady=10,
                                                                                                  sticky="ew")
        ctk.CTkButton(form_frame, text="Clear All Stock", command=self.clear_stock_secure, fg_color="red",
                      hover_color="darkred").grid(row=5, column=0, columnspan=2, padx=10, pady=10, sticky="ew")
        inventory_list_frame = ctk.CTkFrame(self.inventory_frame)
        inventory_list_frame.pack(fill="both", expand=True, padx=20, pady=10)
        self.inventory_tree = ttk.Treeview(inventory_list_frame, columns=("ID", "Name", "Price", "Quantity"),
                                           show='headings')
        self.inventory_tree.heading("ID", text="ID");
        self.inventory_tree.heading("Name", text="Name");
        self.inventory_tree.heading("Price", text="Price (Rs.)");
        self.inventory_tree.heading("Quantity", text="Quantity")
        self.inventory_tree.pack(fill="both", expand=True)
        self.inventory_tree.bind("<<TreeviewSelect>>", self.on_product_select)
        self.refresh_inventory_list()

    def add_image(self):
        path = filedialog.askopenfilename(filetypes=[("Image Files", "*.jpg *.jpeg *.png")])
        if path:
            self.image_path_label.configure(text=os.path.basename(path));
            self.image_path = path

    def refresh_inventory_list(self):
        for item in self.inventory_tree.get_children(): self.inventory_tree.delete(item)
        for row in self.cursor.execute("SELECT id, name, price, quantity FROM products"):
            self.inventory_tree.insert("", "end", values=row)

    def add_product_secure(self):
        if self.ask_password(): self.add_product()

    def add_product(self):
        name, price_str, qty_str = self.product_name_entry.get(), self.product_price_entry.get(), self.product_qty_entry.get()
        image_path = getattr(self, 'image_path', "")
        if not all([name, price_str, qty_str]): return messagebox.showerror("Error", "Please fill out all fields.")
        try:
            price, qty = float(price_str), int(qty_str)
            self.cursor.execute("INSERT INTO products (name, price, quantity, image_path) VALUES (?, ?, ?, ?)",
                                (name, price, qty, image_path))
            self.conn.commit()
            messagebox.showinfo("Success", f"Product '{name}' added.")
            self.clear_inventory_form();
            self.refresh_inventory_list()
        except (ValueError, sqlite3.Error) as e:
            messagebox.showerror("Error", f"Invalid input or database error: {e}")

    def update_product_secure(self):
        if self.ask_password(): self.update_product()

    def update_product(self):
        product_id, name, price_str, qty_str = self.product_id_entry.get(), self.product_name_entry.get(), self.product_price_entry.get(), self.product_qty_entry.get()
        image_path = getattr(self, 'image_path', None)
        if not product_id: return messagebox.showerror("Error", "Please select a product.")
        if not all([name, price_str, qty_str]): return messagebox.showerror("Error", "Please fill out all fields.")
        try:
            price, qty = float(price_str), int(qty_str)
            if image_path is not None:
                self.cursor.execute("UPDATE products SET name=?, price=?, quantity=?, image_path=? WHERE id=?",
                                    (name, price, qty, image_path, product_id))
            else:
                self.cursor.execute("UPDATE products SET name=?, price=?, quantity=? WHERE id=?",
                                    (name, price, qty, product_id))
            self.conn.commit()
            messagebox.showinfo("Success", f"Product ID '{product_id}' updated.")
            self.clear_inventory_form();
            self.refresh_inventory_list()
        except (ValueError, sqlite3.Error) as e:
            messagebox.showerror("Error", f"Invalid input or database error: {e}")

    def clear_stock_secure(self):
        if self.ask_password(): self.clear_stock()

    def clear_stock(self):
        if messagebox.askyesno("Confirm Clear Stock", "Are you sure? This cannot be undone."):
            try:
                self.cursor.execute("DELETE FROM products");
                self.cursor.execute("DELETE FROM sqlite_sequence WHERE name='products'")
                self.conn.commit()
                messagebox.showinfo("Success", "All products cleared.");
                self.refresh_inventory_list()
            except sqlite3.Error as e:
                messagebox.showerror("Database Error", f"An error occurred: {e}")

    def on_product_select(self, event):
        selected_item = self.inventory_tree.focus()
        if not selected_item: return
        values = self.inventory_tree.item(selected_item)['values']
        self.product_id_entry.delete(0, tk.END);
        self.product_id_entry.insert(0, values[0])
        self.product_name_entry.delete(0, tk.END);
        self.product_name_entry.insert(0, values[1])
        self.product_price_entry.delete(0, tk.END);
        self.product_price_entry.insert(0, values[2])
        self.product_qty_entry.delete(0, tk.END);
        self.product_qty_entry.insert(0, values[3])
        path = self.cursor.execute("SELECT image_path FROM products WHERE id=?", (values[0],)).fetchone()[0]
        if path and os.path.exists(path):
            self.image_path = path
            self.image_path_label.configure(text=os.path.basename(path))
            img = ctk.CTkImage(light_image=Image.open(path), size=(100, 100))
            self.product_image_label.configure(image=img, text="")
        else:
            self.product_image_label.configure(image=None, text="No Image")
            self.image_path_label.configure(text="No image selected");
            self.image_path = ""

    def clear_inventory_form(self):
        self.product_id_entry.delete(0, tk.END);
        self.product_name_entry.delete(0, tk.END)
        self.product_price_entry.delete(0, tk.END);
        self.product_qty_entry.delete(0, tk.END)
        self.image_path_label.configure(text="No image selected")
        self.product_image_label.configure(image=None, text="No Image");
        self.image_path = ""

    # --- POS Section ---
    def create_pos_ui(self):
        self.pos_frame.grid_columnconfigure(0, weight=2);
        self.pos_frame.grid_columnconfigure(1, weight=1);
        self.pos_frame.grid_rowconfigure(0, weight=1)
        pos_left_frame = ctk.CTkFrame(self.pos_frame)
        pos_left_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        ctk.CTkLabel(pos_left_frame, text="Products", font=ctk.CTkFont(size=16, weight="bold")).pack(pady=5)
        self.search_entry = ctk.CTkEntry(pos_left_frame, placeholder_text="Search by name...")
        self.search_entry.pack(fill="x", padx=10, pady=5)
        self.search_entry.bind("<KeyRelease>", self.search_product)
        self.product_grid_frame = ctk.CTkScrollableFrame(pos_left_frame, label_text="")
        self.product_grid_frame.pack(fill="both", expand=True, padx=10, pady=10)

        pos_right_frame = ctk.CTkFrame(self.pos_frame)
        pos_right_frame.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)
        ctk.CTkLabel(pos_right_frame, text="Cart", font=ctk.CTkFont(size=16, weight="bold")).pack(pady=5)
        self.cart_tree = ttk.Treeview(pos_right_frame, columns=("ID", "Name", "Price", "Quantity"), show='headings')
        self.cart_tree.heading("ID", text="ID");
        self.cart_tree.heading("Name", text="Name");
        self.cart_tree.heading("Price", text="Price (Rs.)");
        self.cart_tree.heading("Quantity", text="Quantity")
        self.cart_tree.pack(fill="both", expand=True, padx=10, pady=10)

        cart_controls_frame = ctk.CTkFrame(pos_right_frame)
        cart_controls_frame.pack(fill="x", padx=10, pady=(0, 5))
        cart_controls_frame.grid_columnconfigure((0, 1, 2), weight=1)
        ctk.CTkButton(cart_controls_frame, text="+", font=ctk.CTkFont(size=20),
                      command=self.increase_cart_quantity).grid(row=0, column=0, padx=2, sticky="ew")
        ctk.CTkButton(cart_controls_frame, text="-", font=ctk.CTkFont(size=20),
                      command=self.decrease_cart_quantity).grid(row=0, column=1, padx=2, sticky="ew")
        ctk.CTkButton(cart_controls_frame, text="Remove Item", fg_color="red", hover_color="darkred",
                      command=self.remove_from_cart).grid(row=0, column=2, padx=2, sticky="ew")

        self.total_label = ctk.CTkLabel(pos_right_frame, text="Total: Rs.0.00",
                                        font=ctk.CTkFont(size=18, weight="bold"))
        self.total_label.pack(pady=10)

        checkout_frame = ctk.CTkFrame(pos_right_frame)
        checkout_frame.pack(fill="x", padx=10, pady=5)
        checkout_frame.grid_columnconfigure((0, 1), weight=1)
        # --- THIS IS THE CHANGE ---
        ctk.CTkButton(checkout_frame, text="Checkout", command=self.checkout_secure, fg_color="green",
                      hover_color="darkgreen").grid(row=0, column=0, padx=2, sticky="ew")
        self.print_button = ctk.CTkButton(checkout_frame, text="Print Receipt", command=self.print_receipt,
                                          state="disabled")
        self.print_button.grid(row=0, column=1, padx=2, sticky="ew")

        ctk.CTkButton(pos_right_frame, text="Clear Cart", command=self.clear_cart).pack(fill="x", padx=10, pady=5)
        self.cart = {}

    def populate_product_grid(self, search_term=""):
        for widget in self.product_grid_frame.winfo_children(): widget.destroy()
        products = self.cursor.execute(
            "SELECT id, name, price, image_path FROM products WHERE quantity > 0 AND name LIKE ?",
            (f"%{search_term}%",)).fetchall()
        for i, (pid, name, price, path) in enumerate(products):
            row, col = divmod(i, 4)
            item_frame = ctk.CTkFrame(self.product_grid_frame)
            item_frame.grid(row=row, column=col, padx=10, pady=10, sticky="nsew")
            try:
                img = Image.open(path) if path and os.path.exists(path) else Image.new('RGB', (150, 120), color='grey')
                ctk_img = ctk.CTkImage(light_image=img, size=(150, 120))
                img_label = ctk.CTkLabel(item_frame, image=ctk_img, text="")
            except Exception:
                img_label = ctk.CTkLabel(item_frame, text="No Image", width=150, height=120)
            img_label.pack(pady=(10, 5))
            ctk.CTkLabel(item_frame, text=name, font=ctk.CTkFont(size=14, weight="bold")).pack()
            ctk.CTkLabel(item_frame, text=f"Rs.{price:.2f}", font=ctk.CTkFont(size=12)).pack(pady=(0, 10))

            add_func = lambda e, p=pid: self.add_to_cart(p, 1)
            item_frame.bind("<Button-1>", add_func);
            img_label.bind("<Button-1>", add_func)

    def search_product(self, event=None):
        self.populate_product_grid(self.search_entry.get())

    def add_to_cart(self, product_id, quantity):
        name, price, stock = self.cursor.execute("SELECT name, price, quantity FROM products WHERE id=?",
                                                 (product_id,)).fetchone()
        cart_qty = self.cart.get(product_id, {}).get('quantity', 0)
        if (quantity + cart_qty) > stock: return messagebox.showwarning("Out of Stock",
                                                                        f"Only {stock - cart_qty} more available.")
        if product_id in self.cart:
            self.cart[product_id]['quantity'] += quantity
        else:
            self.cart[product_id] = {'name': name, 'price': price, 'quantity': quantity}
        self.refresh_cart_tree()

    def get_selected_cart_product_id(self):
        selected_item = self.cart_tree.focus()
        return int(self.cart_tree.item(selected_item)['values'][0]) if selected_item else None

    def increase_cart_quantity(self):
        if pid := self.get_selected_cart_product_id(): self.add_to_cart(pid, 1)

    def decrease_cart_quantity(self):
        pid = self.get_selected_cart_product_id()
        if not pid: return
        if self.cart[pid]['quantity'] > 1:
            self.cart[pid]['quantity'] -= 1
        else:
            del self.cart[pid]
        self.refresh_cart_tree()

    def remove_from_cart(self):
        if pid := self.get_selected_cart_product_id():
            if pid in self.cart: del self.cart[pid]; self.refresh_cart_tree()

    def refresh_cart_tree(self):
        for item in self.cart_tree.get_children(): self.cart_tree.delete(item)
        total = sum(item['price'] * item['quantity'] for item in self.cart.values())
        self.total_label.configure(text=f"Total: Rs.{total:.2f}")
        for pid, item in self.cart.items():
            self.cart_tree.insert("", "end", values=(pid, item['name'], item['price'], item['quantity']))
        self.print_button.configure(state="disabled")

    def clear_cart(self):
        self.cart = {};
        self.refresh_cart_tree()

    def checkout_secure(self):
        if self.ask_password(): self.checkout()

    def checkout(self):
        if not self.cart: return messagebox.showerror("Error", "Cart is empty.")
        total_price = sum(item['price'] * item['quantity'] for item in self.cart.values())
        try:
            self.cursor.execute("INSERT INTO sales (total_price) VALUES (?)", (total_price,))
            sale_id = self.cursor.lastrowid
            self.last_sale_details = {"sale_id": sale_id, "items": [], "total": total_price,
                                      "date": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
            for pid, item in self.cart.items():
                self.cursor.execute(
                    "INSERT INTO sale_items (sale_id, product_id, product_name, quantity, price) VALUES (?, ?, ?, ?, ?)",
                    (sale_id, pid, item['name'], item['quantity'], item['price']))
                self.cursor.execute("UPDATE products SET quantity = quantity - ? WHERE id = ?", (item['quantity'], pid))
                self.last_sale_details["items"].append(item)
            self.conn.commit()
            messagebox.showinfo("Success", "Checkout complete.")
            self.clear_cart();
            self.populate_product_grid()
            self.print_button.configure(state="normal")
        except sqlite3.Error as e:
            self.conn.rollback();
            messagebox.showerror("Database Error", f"Checkout failed: {e}")

    def print_receipt(self):
        if not self.last_sale_details:
            return messagebox.showerror("Error", "No sale has been made yet.")
        receipt = f"""
        SAMARA TRADE CENTER
        --------------------------
        Sale ID: {self.last_sale_details['sale_id']}
        Date: {self.last_sale_details['date']}
        --------------------------
        Items:
        """
        for item in self.last_sale_details['items']:
            receipt += f"\n  {item['name']} x {item['quantity']}"
            receipt += f"\n    (Rs.{item['price']:.2f} each) = Rs.{item['price'] * item['quantity']:.2f}"

        receipt += f"""
        --------------------------
        Total: Rs.{self.last_sale_details['total']:.2f}

        Thank you!
        """

        try:
            printer_name = win32print.GetDefaultPrinter()
            hPrinter = win32print.OpenPrinter(printer_name)
            try:
                hJob = win32print.StartDocPrinter(hPrinter, 1, ("Sale Receipt", None, "RAW"))
                try:
                    win32print.StartPagePrinter(hPrinter)
                    win32print.WritePrinter(hPrinter, receipt.encode())
                    win32print.EndPagePrinter(hPrinter)
                finally:
                    win32print.EndDocPrinter(hPrinter)
            finally:
                win32print.ClosePrinter(hPrinter)
            messagebox.showinfo("Success", "Receipt sent to printer.")
        except Exception as e:
            messagebox.showerror("Printing Error", f"Could not print receipt: {e}")

    # --- Sales Section ---
    def create_sales_ui(self):
        sales_main_frame = ctk.CTkFrame(self.sales_frame)
        sales_main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        sales_main_frame.grid_columnconfigure(1, weight=1);
        sales_main_frame.grid_rowconfigure(0, weight=1)
        sales_list_frame = ctk.CTkFrame(sales_main_frame)
        sales_list_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 10));
        sales_list_frame.grid_rowconfigure(1, weight=1)
        controls_frame = ctk.CTkFrame(sales_list_frame);
        controls_frame.pack(fill="x", pady=10)
        ctk.CTkLabel(controls_frame, text="Sales History", font=ctk.CTkFont(size=16, weight="bold")).pack(side="left",
                                                                                                          padx=10)
        ctk.CTkButton(controls_frame, text="Clear All Sales", command=self.clear_sales_secure, fg_color="red",
                      hover_color="darkred").pack(side="right", padx=10)
        self.sales_tree = ttk.Treeview(sales_list_frame, columns=("ID", "Total", "Date"), show='headings')
        self.sales_tree.heading("ID", text="Sale ID");
        self.sales_tree.heading("Total", text="Total (Rs.)");
        self.sales_tree.heading("Date", text="Date")
        self.sales_tree.pack(fill="both", expand=True, padx=10, pady=10)
        self.sales_tree.bind("<<TreeviewSelect>>", self.on_sale_select)
        sale_details_frame = ctk.CTkFrame(sales_main_frame)
        sale_details_frame.grid(row=0, column=1, sticky="nsew");
        sale_details_frame.grid_rowconfigure(1, weight=1)
        ctk.CTkLabel(sale_details_frame, text="Sale Details", font=ctk.CTkFont(size=16, weight="bold")).pack(pady=10)
        self.sale_items_tree = ttk.Treeview(sale_details_frame, columns=("Product", "Qty", "Price"), show='headings')
        self.sale_items_tree.heading("Product", text="Product");
        self.sale_items_tree.heading("Qty", text="Quantity");
        self.sale_items_tree.heading("Price", text="Price (Rs.)")
        self.sale_items_tree.pack(fill="both", expand=True, padx=10, pady=10)
        self.refresh_sales_list()

    def refresh_sales_list(self):
        for item in self.sales_tree.get_children(): self.sales_tree.delete(item)
        for row in self.cursor.execute("SELECT id, total_price, sale_date FROM sales ORDER BY sale_date DESC"):
            self.sales_tree.insert("", "end", values=row)

    def on_sale_select(self, event):
        selected_item = self.sales_tree.focus()
        if not selected_item: return
        sale_id = self.sales_tree.item(selected_item)['values'][0]
        for item in self.sale_items_tree.get_children(): self.sale_items_tree.delete(item)
        for row in self.cursor.execute("SELECT product_name, quantity, price FROM sale_items WHERE sale_id=?",
                                       (sale_id,)):
            self.sale_items_tree.insert("", "end", values=row)

    def clear_sales_secure(self):
        if self.ask_password(): self.clear_sales()

    def clear_sales(self):
        if messagebox.askyesno("Confirm Clear Sales", "Are you sure? This cannot be undone."):
            try:
                self.cursor.execute("DELETE FROM sales");
                self.cursor.execute("DELETE FROM sale_items")
                self.cursor.execute("DELETE FROM sqlite_sequence WHERE name IN ('sales', 'sale_items')")
                self.conn.commit()
                messagebox.showinfo("Success", "All sales history has been cleared.")
                self.refresh_sales_list();
                [self.sale_items_tree.delete(i) for i in self.sale_items_tree.get_children()]
            except sqlite3.Error as e:
                messagebox.showerror("Database Error", f"An error occurred: {e}")

    # --- Export Section ---
    def export_to_excel_secure(self):
        if self.ask_password(): self.export_to_excel()

    def export_to_excel(self):
        filename = filedialog.asksaveasfilename(
            initialfile=f"retail_export_{datetime.now().strftime('%Y%m%d')}.xlsx",
            defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if not filename: return
        try:
            wb = Workbook()
            ws_products = wb.active;
            ws_products.title = "Products"
            ws_products.append(["ID", "Name", "Price (Rs.)", "Quantity", "Image Path"])
            for row in self.cursor.execute(
                "SELECT id, name, price, quantity, image_path FROM products"): ws_products.append(row)
            ws_sales = wb.create_sheet(title="Recent Sales")
            ws_sales.append(["Sale ID", "Total Price (Rs.)", "Date"])
            today_sales = self.cursor.execute(
                "SELECT id, total_price, sale_date FROM sales WHERE date(sale_date) = date('now')").fetchall()
            sale_ids = [row[0] for row in today_sales]
            for row in today_sales: ws_sales.append(row)
            ws_sale_items = wb.create_sheet(title="Recent Sale Items")
            ws_sale_items.append(["Sale Item ID", "Sale ID", "Product ID", "Product Name", "Quantity", "Price (Rs.)"])
            if sale_ids:
                placeholders = ','.join('?' * len(sale_ids))
                query = f"SELECT id, sale_id, product_id, product_name, quantity, price FROM sale_items WHERE sale_id IN ({placeholders})"
                for row in self.cursor.execute(query, sale_ids): ws_sale_items.append(row)
            wb.save(filename)
            messagebox.showinfo("Success", f"Data exported to {filename}")
        except Exception as e:
            messagebox.showerror("Export Error", f"An error occurred during export: {e}")

    # --- Settings Section ---
    def create_settings_ui(self):
        settings_main_frame = ctk.CTkFrame(self.settings_frame)
        settings_main_frame.pack(padx=20, pady=20, fill="x")
        ctk.CTkLabel(settings_main_frame, text="Change Password", font=ctk.CTkFont(size=18, weight="bold")).pack(
            pady=(0, 10))
        self.old_password_entry = ctk.CTkEntry(settings_main_frame, placeholder_text="Old Password", show="*")
        self.old_password_entry.pack(fill="x", padx=20, pady=5)
        self.new_password_entry = ctk.CTkEntry(settings_main_frame, placeholder_text="New Password", show="*")
        self.new_password_entry.pack(fill="x", padx=20, pady=5)
        self.confirm_password_entry = ctk.CTkEntry(settings_main_frame, placeholder_text="Confirm New Password",
                                                   show="*")
        self.confirm_password_entry.pack(fill="x", padx=20, pady=5)
        ctk.CTkButton(settings_main_frame, text="Save New Password", command=self.change_password).pack(padx=20,
                                                                                                        pady=20)

    def change_password(self):
        old_password, new_password, confirm_password = self.old_password_entry.get(), self.new_password_entry.get(), self.confirm_password_entry.get()
        if not self.verify_password(old_password): return messagebox.showerror("Error", "Old password is not correct.")
        if not new_password: return messagebox.showerror("Error", "New password cannot be empty.")
        if new_password != confirm_password: return messagebox.showerror("Error", "New passwords do not match.")
        try:
            self.cursor.execute("UPDATE settings SET password = ? WHERE id = 1", (self.hash_password(new_password),))
            self.conn.commit()
            messagebox.showinfo("Success", "Password changed successfully.")
            self.old_password_entry.delete(0, tk.END);
            self.new_password_entry.delete(0, tk.END);
            self.confirm_password_entry.delete(0, tk.END)
        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"An error occurred: {e}")

    # --- Frame Navigation ---
    def select_frame_by_name(self, name):
        buttons = {"inventory": self.inventory_button, "pos": self.pos_button, "sales": self.sales_button,
                   "settings": self.settings_button}
        frames = {"inventory": self.inventory_frame, "pos": self.pos_frame, "sales": self.sales_frame,
                  "settings": self.settings_frame}
        for frame_name, button in buttons.items():
            button.configure(fg_color=("gray75", "gray25") if name == frame_name else "transparent")
        for frame in frames.values():
            frame.grid_forget()
        frames[name].grid(row=0, column=1, sticky="nsew", padx=20, pady=20)
        if name == "inventory":
            self.refresh_inventory_list()
        elif name == "pos":
            self.populate_product_grid()
        elif name == "sales":
            self.refresh_sales_list()

    def inventory_button_event(self):
        self.select_frame_by_name("inventory")

    def pos_button_event(self):
        self.select_frame_by_name("pos")

    def sales_button_event(self):
        self.select_frame_by_name("sales")

    def settings_button_event(self):
        self.select_frame_by_name("settings")


if __name__ == "__main__":
    app = RetailApp()
    app.mainloop()
    app.conn.close()
